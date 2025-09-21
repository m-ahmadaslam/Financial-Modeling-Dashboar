from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json
from typing import Dict, List, Any, Optional
from pydantic import BaseModel
import os
from datetime import datetime
import re

app = FastAPI(
    title="Financial Dashboard Backend",
    description="FastAPI backend for Excel parsing and formula analysis",
    version="1.0.0"
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://127.0.0.1:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Pydantic models
class FieldInfo(BaseModel):
    id: str
    name: str
    row: int
    column: str
    type: str  # 'input' or 'calculated'
    dataType: str  # 'text', 'number', 'currency', 'percentage', 'date', 'boolean'
    value: Any
    formula: Optional[str] = None
    isNamedCell: bool = False
    namedCell: Optional[str] = None
    required: bool = False
    unit: Optional[str] = None
    section: str
    heading: str

class SectionInfo(BaseModel):
    id: str
    name: str
    row: int
    headings: Dict[str, Any]
    fields: List[FieldInfo]

class ExcelAnalysisResult(BaseModel):
    totalSheets: int
    totalFields: int
    inputFields: int
    calculatedFields: int
    sections: List[SectionInfo]
    formulaPatterns: Dict[str, int]
    analysisTimestamp: str

class FormulaPattern(BaseModel):
    pattern: str
    count: int
    examples: List[str]

# Global variables for caching
excel_data_cache = {}
current_analysis = None

@app.get("/")
async def root():
    return {"message": "Financial Dashboard Backend API", "status": "running"}

@app.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}

@app.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    """Upload and analyze Excel file"""
    try:
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Only Excel files are allowed")
        
        # Read file content
        content = await file.read()
        
        # Load workbook
        workbook = load_workbook(filename=file.filename, data_only=False)
        
        # Analyze all sheets
        analysis_result = analyze_excel_workbook(workbook, file.filename)
        
        # Cache the result
        global current_analysis
        current_analysis = analysis_result
        
        return analysis_result
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing Excel file: {str(e)}")

@app.get("/analyze-excel/{filename}")
async def analyze_excel_file(filename: str):
    """Analyze Excel file from project root"""
    try:
        file_path = f"../{filename}"
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="Excel file not found")
        
        # Load workbook
        workbook = load_workbook(filename=file_path, data_only=False)
        
        # Analyze all sheets
        analysis_result = analyze_excel_workbook(workbook, filename)
        
        # Cache the result
        global current_analysis
        current_analysis = analysis_result
        
        return analysis_result
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error analyzing Excel file: {str(e)}")

@app.get("/get-analysis")
async def get_current_analysis():
    """Get the current analysis result"""
    if current_analysis is None:
        raise HTTPException(status_code=404, detail="No analysis available")
    return current_analysis

@app.get("/export-csv")
async def export_analysis_to_csv():
    """Export current analysis to CSV format"""
    if current_analysis is None:
        raise HTTPException(status_code=404, detail="No analysis available")
    
    try:
        # Convert to CSV format
        csv_data = convert_analysis_to_csv(current_analysis)
        return JSONResponse(content=csv_data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting to CSV: {str(e)}")

def analyze_excel_workbook(workbook, filename: str) -> ExcelAnalysisResult:
    """Comprehensive Excel workbook analysis"""
    print(f"🔍 Analyzing Excel file: {filename}")
    
    all_sections = []
    total_fields = 0
    input_fields = 0
    calculated_fields = 0
    formula_patterns = {}
    
    # Analyze each sheet
    for sheet_name in workbook.sheetnames:
        print(f"📊 Analyzing sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        
        # Extract sections and fields from this sheet
        sheet_sections = extract_sections_from_sheet(sheet, sheet_name)
        
        for section in sheet_sections:
            all_sections.append(section)
            total_fields += len(section.fields)
            
            for field in section.fields:
                if field.type == 'input':
                    input_fields += 1
                elif field.type == 'calculated':
                    calculated_fields += 1
                
                # Analyze formula patterns
                if field.formula:
                    pattern = analyze_formula_pattern(field.formula)
                    if pattern in formula_patterns:
                        formula_patterns[pattern] += 1
                    else:
                        formula_patterns[pattern] = 1
    
    return ExcelAnalysisResult(
        totalSheets=len(workbook.sheetnames),
        totalFields=total_fields,
        inputFields=input_fields,
        calculatedFields=calculated_fields,
        sections=all_sections,
        formulaPatterns=formula_patterns,
        analysisTimestamp=datetime.now().isoformat()
    )

def extract_sections_from_sheet(sheet, sheet_name: str) -> List[SectionInfo]:
    """Extract sections and fields from a single sheet"""
    sections = []
    current_section = None
    current_heading = None
    section_id = 1
    
    # Get sheet dimensions
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    print(f"   Sheet dimensions: {max_row} rows x {max_col} columns")
    
    # Scan through all rows
    for row_num in range(1, max_row + 1):
        row_data = []
        for col_num in range(1, max_col + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            row_data.append({
                'value': cell.value,
                'formula': cell.data_type == 'f' and str(cell.value) or None,
                'coordinate': cell.coordinate,
                'row': row_num,
                'column': col_num
            })
        
        # Check for section headers (blue cells or bold text)
        section_header = detect_section_header(row_data, row_num)
        if section_header:
            # Save previous section if exists
            if current_section:
                sections.append(current_section)
            
            # Start new section
            current_section = SectionInfo(
                id=f"section_{section_id}",
                name=section_header,
                row=row_num,
                headings={},
                fields=[]
            )
            section_id += 1
            current_heading = None
            continue
        
        # Check for heading within current section
        if current_section:
            heading = detect_heading(row_data, row_num)
            if heading:
                current_heading = heading
                current_section.headings[heading] = {
                    'id': heading.lower().replace(' ', '_'),
                    'name': heading,
                    'fields': []
                }
                continue
            
            # Check for field definitions
            field_info = detect_field(row_data, row_num, current_section.id, current_heading)
            if field_info:
                current_section.fields.append(field_info)
                if current_heading and current_heading in current_section.headings:
                    current_section.headings[current_heading]['fields'].append(field_info)
    
    # Add the last section
    if current_section:
        sections.append(current_section)
    
    print(f"   Found {len(sections)} sections in sheet '{sheet_name}'")
    return sections

def detect_section_header(row_data: List[Dict], row_num: int) -> Optional[str]:
    """Detect if this row contains a section header"""
    for cell in row_data:
        if cell['value'] and isinstance(cell['value'], str):
            value = cell['value'].strip()
            # Look for section-like headers - more specific to financial models
            if (len(value) > 3 and 
                not value.isdigit() and 
                not value.startswith('=') and
                not value.startswith('F') and
                (value.isupper() or 
                 any(keyword in value.lower() for keyword in [
                     'project', 'cost', 'debt', 'equity', 'revenue', 'tax', 'capacity', 
                     'technical', 'operation', 'construction', 'financing', 'macroeconomic',
                     'assumption', 'input', 'output', 'calculation', 'timeline', 'sponsor',
                     'plant', 'tariff', 'sensitivity', 'structure', 'working', 'capital',
                     'performance', 'service', 'accounting', 'liquidated', 'damages'
                 ]))):
                return value
    return None

def detect_heading(row_data: List[Dict], row_num: int) -> Optional[str]:
    """Detect if this row contains a heading"""
    for cell in row_data:
        if cell['value'] and isinstance(cell['value'], str):
            value = cell['value'].strip()
            # Look for heading-like text
            if (len(value) > 2 and 
                not value.isdigit() and 
                not value.startswith('=') and
                not value.startswith('F') and
                len(value.split()) <= 5):  # Short phrases
                return value
    return None

def detect_field(row_data: List[Dict], row_num: int, section_id: str, heading: str) -> Optional[FieldInfo]:
    """Detect if this row contains a field definition"""
    for col_idx, cell in enumerate(row_data):
        if cell['value'] is not None:
            # Look for formulas or field names
            if cell['formula'] or (isinstance(cell['value'], str) and len(cell['value']) > 2):
                # Try to find field name in nearby cells
                field_name = find_field_name(row_data, col_idx)
                if field_name:
                    # Determine field type
                    field_type = 'calculated' if cell['formula'] else 'input'
                    
                    # Determine data type
                    data_type = determine_data_type(cell['value'])
                    
                    return FieldInfo(
                        id=f"field_{row_num}",
                        name=field_name,
                        row=row_num,
                        column=cell['coordinate'],
                        type=field_type,
                        dataType=data_type,
                        value=cell['value'],
                        formula=cell['formula'],
                        isNamedCell=bool(cell['formula'] and 'INDEX' in str(cell['formula'])),
                        namedCell=cell['coordinate'] if cell['formula'] and 'INDEX' in str(cell['formula']) else None,
                        required=False,
                        section=section_id,
                        heading=heading or 'general'
                    )
    return None

def find_field_name(row_data: List[Dict], col_idx: int) -> Optional[str]:
    """Find field name in nearby cells"""
    # Check cells to the left first
    for offset in range(1, 4):
        check_idx = col_idx - offset
        if check_idx >= 0 and row_data[check_idx]['value']:
            value = str(row_data[check_idx]['value']).strip()
            if (len(value) > 2 and 
                not value.isdigit() and 
                not value.startswith('=') and
                not value.startswith('F')):
                return value
    
    # Check cells to the right
    for offset in range(1, 3):
        check_idx = col_idx + offset
        if check_idx < len(row_data) and row_data[check_idx]['value']:
            value = str(row_data[check_idx]['value']).strip()
            if (len(value) > 2 and 
                not value.isdigit() and 
                not value.startswith('=') and
                not value.startswith('F')):
                return value
    
    return None

def determine_data_type(value: Any) -> str:
    """Determine the data type of a field value"""
    if value is None:
        return 'text'
    
    if isinstance(value, (int, float)):
        return 'number'
    
    if isinstance(value, str):
        if value.lower() in ['true', 'false']:
            return 'boolean'
        if '%' in value:
            return 'percentage'
        if '$' in value or 'USD' in value or 'SAR' in value:
            return 'currency'
        if re.match(r'\d{4}-\d{2}-\d{2}', value):
            return 'date'
    
    return 'text'

def analyze_formula_pattern(formula: str) -> str:
    """Analyze formula and return pattern type"""
    if 'INDEX' in formula:
        return 'INDEX'
    elif 'EDATE' in formula:
        return 'EDATE'
    elif 'IF(' in formula:
        return 'IF'
    elif re.search(r'[+\-*/]', formula):
        return 'ARITHMETIC'
    elif re.search(r'F\d+', formula):
        return 'CELL_REF'
    else:
        return 'UNKNOWN'

def convert_analysis_to_csv(analysis: ExcelAnalysisResult) -> Dict[str, Any]:
    """Convert analysis result to CSV-friendly format"""
    csv_data = {
        'metadata': {
            'totalSheets': analysis.totalSheets,
            'totalFields': analysis.totalFields,
            'inputFields': analysis.inputFields,
            'calculatedFields': analysis.calculatedFields,
            'analysisTimestamp': analysis.analysisTimestamp
        },
        'sections': [],
        'fields': []
    }
    
    for section in analysis.sections:
        section_data = {
            'id': section.id,
            'name': section.name,
            'row': section.row,
            'fieldCount': len(section.fields)
        }
        csv_data['sections'].append(section_data)
        
        for field in section.fields:
            field_data = {
                'id': field.id,
                'name': field.name,
                'row': field.row,
                'column': field.column,
                'type': field.type,
                'dataType': field.dataType,
                'value': field.value,
                'formula': field.formula,
                'isNamedCell': field.isNamedCell,
                'namedCell': field.namedCell,
                'required': field.required,
                'unit': field.unit,
                'section': field.section,
                'heading': field.heading
            }
            csv_data['fields'].append(field_data)
    
    return csv_data

# Timeline Generation Models
class TimelineInputs(BaseModel):
    model_start_date: str
    ppa_signing_date: str
    financial_close_date: Optional[str] = None
    construction_period_start_date: Optional[str] = None
    construction_period: int  # months
    scheduled_pcod_as_per_ppa: str
    scheduled_pcod: str
    commercial_operation_date: Optional[str] = None
    tenor_of_ppa: int  # years
    end_of_commercial_operations: str
    extension_in_ppa: int  # years
    end_of_extension_period: str
    months_in_quarterly_period: int

class TimelineResponse(BaseModel):
    monthly: Dict[str, Any]
    quarterly: Dict[str, Any]
    semiannual: Dict[str, Any]
    annual: Dict[str, Any]
    metadata: Dict[str, Any]

@app.post("/generate-timelines", response_model=TimelineResponse)
async def generate_timelines(inputs: TimelineInputs):
    """Generate comprehensive project timelines"""
    try:
        print(f"🕒 Generating timelines for project starting {inputs.model_start_date}")
        
        # Generate all timeline types
        monthly_timeline = generate_monthly_timeline(inputs)
        quarterly_timeline = generate_quarterly_timeline(inputs)
        semiannual_timeline = generate_semiannual_timeline(inputs)
        annual_timeline = generate_annual_timeline(inputs)
        
        return TimelineResponse(
            monthly=monthly_timeline,
            quarterly=quarterly_timeline,
            semiannual=semiannual_timeline,
            annual=annual_timeline,
            metadata={
                "generation_timestamp": datetime.now().isoformat(),
                "project_start": inputs.model_start_date,
                "project_end": inputs.end_of_extension_period,
                "total_periods": {
                    "monthly": monthly_timeline.get("total_periods", 0),
                    "quarterly": quarterly_timeline.get("total_periods", 0),
                    "semiannual": semiannual_timeline.get("total_periods", 0),
                    "annual": annual_timeline.get("total_periods", 0)
                }
            }
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Timeline generation failed: {str(e)}")

def generate_monthly_timeline(inputs: TimelineInputs) -> Dict[str, Any]:
    """Generate monthly timeline with 200 rows × 485+ columns"""
    import pandas as pd
    from datetime import datetime, timedelta
    from dateutil.relativedelta import relativedelta
    
    start_date = datetime.strptime(inputs.model_start_date, '%Y-%m-%d')
    end_date = datetime.strptime(inputs.end_of_extension_period, '%Y-%m-%d')
    
    # Generate monthly periods (485+ columns as shown in Excel)
    periods = []
    current_date = start_date
    period_counter = 1
    
    while current_date <= end_date:
        month_start = current_date.replace(day=1)
        month_end = (month_start + relativedelta(months=1)) - timedelta(days=1)
        
        periods.append({
            'period': period_counter,
            'date': current_date.strftime('%d-%b-%y'),
            'month_start': month_start.strftime('%d-%b-%y'),
            'month_end': month_end.strftime('%d-%b-%y'),
            'days_in_month': month_end.day,
            'financial_year': get_financial_year(current_date),
            'project_year': get_project_year(current_date, start_date),
            'quarter': get_quarter(current_date),
            'year': current_date.year
        })
        
        current_date += relativedelta(months=1)
        period_counter += 1
    
    # Generate timeline rows (200+ rows as shown in Excel)
    timeline_rows = generate_timeline_rows(inputs, periods)
    
    return {
        "type": "monthly",
        "total_periods": len(periods),
        "start_date": inputs.model_start_date,
        "end_date": inputs.end_of_extension_period,
        "columns": periods,
        "rows": timeline_rows
    }

def generate_timeline_rows(inputs: TimelineInputs, periods: List[Dict]) -> List[Dict]:
    """Generate all timeline calculation rows from extracted Excel fields"""
    
    # Load extracted timeline fields
    import json
    try:
        with open('extracted_timeline_fields.json', 'r') as f:
            extracted_fields = json.load(f)
    except:
        # Fallback to basic fields if file not found
        extracted_fields = []
    
    rows = []
    
    # Convert extracted fields to our format
    for field in extracted_fields:
        field_data = {
            "field_name": field.get('name', 'Unknown Field'),
            "type": field.get('type', 'calculated'),
            "unit": field.get('unit', 'text'),
            "formula": field.get('formula', '').replace('= ', '') if field.get('formula') else 'excel_calculation',
            "row": field.get('row', 0),
            "excel_formula": field.get('formula', '')
        }
        
        # Calculate values for this field across all periods
        values = calculate_comprehensive_field_values(field_data, inputs, periods)
        rows.append({
            **field_data,
            "values": values
        })
    
    # If no extracted fields, use basic fallback
    if not extracted_fields:
        basic_fields = [
            {"field_name": "Monthly period", "type": "calculated", "unit": "number", "formula": "sequence"},
            {"field_name": "Month start date", "type": "calculated", "unit": "date", "formula": "month_start"},
            {"field_name": "Month end date", "type": "calculated", "unit": "date", "formula": "month_end"},
            {"field_name": "Days in month", "type": "calculated", "unit": "days", "formula": "days_in_month"},
            {"field_name": "Financial year", "type": "calculated", "unit": "number", "formula": "financial_year"},
        ]
        
        for field in basic_fields:
            values = calculate_field_values(field, inputs, periods)
            rows.append({
                **field,
                "values": values
            })
    
    return rows

def calculate_field_values(field: Dict, inputs: TimelineInputs, periods: List[Dict]) -> List[Any]:
    """Calculate values for a specific field across all periods"""
    values = []
    
    for i, period in enumerate(periods):
        if field["formula"] == "sequence":
            values.append(period["period"])
        elif field["formula"] == "month_start":
            values.append(period["month_start"])
        elif field["formula"] == "month_end":
            values.append(period["month_end"])
        elif field["formula"] == "days_in_month":
            values.append(period["days_in_month"])
        elif field["formula"] == "financial_year":
            values.append(period["financial_year"])
        elif field["formula"] == "project_year":
            values.append(period["project_year"])
        elif field["formula"] == "calendar_year":
            values.append(period["year"])
        elif field["formula"] == "construction_start_flag":
            # Check if this is the construction start month
            construction_start = datetime.strptime(inputs.model_start_date, '%Y-%m-%d')
            period_date = datetime.strptime(period["date"], '%d-%b-%y')
            values.append("#REF!" if period_date.year == construction_start.year and period_date.month == construction_start.month else "#REF!")
        elif field["formula"] == "construction_period_flag":
            # Check if this period is within construction period
            values.append("#REF!")  # Will be calculated based on construction start/end dates
        elif field["formula"] == "commercial_operation_flag":
            # Check if this is the commercial operation start month
            values.append("#REF!")  # Will be calculated based on COD
        else:
            # Default calculation or placeholder
            values.append("#REF!" if field["type"] == "calculated" else 0)
    
    return values

def calculate_comprehensive_field_values(field: Dict, inputs: TimelineInputs, periods: List[Dict]) -> List[Any]:
    """Calculate values for comprehensive timeline fields from Excel"""
    values = []
    field_name = field.get('field_name', '').lower()
    excel_formula = field.get('excel_formula', '')
    
    for i, period in enumerate(periods):
        if field_name == 'monthly period' or 'period' in field_name:
            values.append(period["period"])
        elif 'month start' in field_name:
            values.append(period["month_start"])
        elif 'month end' in field_name:
            values.append(period["month_end"])
        elif 'days in month' in field_name:
            values.append(period["days_in_month"])
        elif 'financial year' in field_name:
            values.append(period["financial_year"])
        elif 'project year' in field_name:
            values.append(period["project_year"])
        elif 'calendar' in field_name or 'calender' in field_name:
            values.append(period["year"])
        elif 'construction start' in field_name:
            # Construction start flag
            values.append(1 if i == 0 else 0)
        elif 'construction end' in field_name:
            # Construction end flag - approximate
            construction_months = inputs.construction_period
            values.append(1 if i == construction_months - 1 else 0)
        elif 'construction period' in field_name or 'construction month flag' in field_name:
            # Construction period flag
            construction_months = inputs.construction_period
            values.append(1 if i < construction_months else 0)
        elif 'construction month counter' in field_name:
            construction_months = inputs.construction_period
            values.append(i + 1 if i < construction_months else 0)
        elif 'commercial operation' in field_name:
            # Commercial operation flag
            construction_months = inputs.construction_period
            values.append(1 if i == construction_months else 0)
        elif 'operation period' in field_name:
            # Operation period flag
            construction_months = inputs.construction_period
            values.append(1 if i >= construction_months else 0)
        elif 'operation month counter' in field_name:
            construction_months = inputs.construction_period
            values.append(max(0, i - construction_months + 1) if i >= construction_months else 0)
        elif 'ppa period' in field_name:
            # PPA period flag
            construction_months = inputs.construction_period
            ppa_months = inputs.tenor_of_ppa * 12
            values.append(1 if construction_months <= i < construction_months + ppa_months else 0)
        elif 'ppa month counter' in field_name:
            construction_months = inputs.construction_period
            ppa_months = inputs.tenor_of_ppa * 12
            if construction_months <= i < construction_months + ppa_months:
                values.append(i - construction_months + 1)
            else:
                values.append(0)
        elif 'debt service' in field_name:
            # Debt service flag - typically starts after construction
            construction_months = inputs.construction_period
            values.append(1 if i >= construction_months else 0)
        elif 'tax' in field_name and 'flag' in field_name:
            # Tax payment flag - typically after operations start
            construction_months = inputs.construction_period
            values.append(1 if i >= construction_months + 12 else 0)  # Start tax after 1 year of operations
        elif 'dividend' in field_name and 'flag' in field_name:
            # Dividend payment flag - typically after operations start
            construction_months = inputs.construction_period
            values.append(1 if i >= construction_months + 6 else 0)  # Start dividends after 6 months of operations
        elif 'reporting period' in field_name:
            # Quarterly reporting periods
            values.append((i // 3) + 1)
        elif excel_formula and 'EDATE' in excel_formula:
            # Date calculation
            values.append(period["month_end"])
        elif excel_formula and 'IF(' in excel_formula:
            # Conditional calculation
            values.append(1 if i % 2 == 0 else 0)
        elif field.get('type') == 'input':
            # Input fields - use default values
            values.append(0)
        else:
            # Default calculated value
            values.append(period.get("period", i + 1))
    
    return values

def get_financial_year(date: datetime, fy_end: str = "March") -> int:
    """Get financial year based on financial year end"""
    if fy_end == "March":
        return date.year + 1 if date.month > 3 else date.year
    else:  # December
        return date.year

def get_project_year(date: datetime, start_date: datetime) -> int:
    """Get project year counter"""
    diff = date - start_date
    return (diff.days // 365) + 1

def get_quarter(date: datetime) -> int:
    """Get quarter number"""
    return (date.month - 1) // 3 + 1

def generate_quarterly_timeline(inputs: TimelineInputs) -> Dict[str, Any]:
    """Generate quarterly timeline"""
    # Simplified quarterly timeline generation
    return {
        "type": "quarterly",
        "total_periods": (inputs.tenor_of_ppa * 4),
        "start_date": inputs.model_start_date,
        "end_date": inputs.end_of_extension_period,
        "columns": [],  # Will be populated with quarterly periods
        "rows": []      # Will be populated with quarterly calculations
    }

def generate_semiannual_timeline(inputs: TimelineInputs) -> Dict[str, Any]:
    """Generate semi-annual timeline"""
    return {
        "type": "semiannual", 
        "total_periods": (inputs.tenor_of_ppa * 2),
        "start_date": inputs.model_start_date,
        "end_date": inputs.end_of_extension_period,
        "columns": [],
        "rows": []
    }

def generate_annual_timeline(inputs: TimelineInputs) -> Dict[str, Any]:
    """Generate annual timeline"""
    return {
        "type": "annual",
        "total_periods": inputs.tenor_of_ppa,
        "start_date": inputs.model_start_date,
        "end_date": inputs.end_of_extension_period,
        "columns": [],
        "rows": []
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
